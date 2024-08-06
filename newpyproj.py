from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
data ={
    20:{
        "Name": "Hak",
        "Age":20
    },
    30:{
        "Name": "Dog",
        "Age":22
    },
    40:{
        "Name": "Chap",
        "Age":24
    }
}


wb = load_workbook('tim.xlsx')
ws = wb['data2']


headings = ['ID'] + list(data[20].keys())
print(list(data[20].values()))
ws.append(headings)

for i in data:
    x = list(data[i].values())
    ws.append([i] + x)
wb.save('tim.xlsx')