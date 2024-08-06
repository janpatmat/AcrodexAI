#Create a workbook or load a workbook
from openpyxl import Workbook, load_workbook

#Path of the worksheet within the folder
# YOU CANNOT SAVE A WORKBOOK IF THE EXCEL IS ALREADY OPENED
wb = load_workbook('Book1.xlsx')
# active worksheet of the workbook
ws = wb.active
ws2 = wb['Sheet2']

wb.create_sheet("Newsheet")

print(wb.sheetnames)

